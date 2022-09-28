from PyQt5.QtWidgets import QApplication, QWidget, QDesktopWidget, QVBoxLayout, QStackedWidget, QLabel, QPushButton, QGraphicsDropShadowEffect, QLineEdit, QScrollArea, QDialog, qApp
from PyQt5.QtGui import QFont, QFontDatabase, QIcon, QPainter
from PyQt5.QtCore import Qt, QTime, QTimer
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from PIL.ImageQt import ImageQt
from pdf2image import convert_from_path
import tempfile
import openpyxl
import docx
import data
import time
import sys

### userData 정의
# [0]: 이름 
# [1]: 학번
# [2]: 동아리명
# [3]: 학과/부
# [4]: 생일
# [5]: 매체

class LoginWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def clearInputFields(self):
        global g_histories
        global g_userData

        self.textField1.setText("")
        self.textField2.setText("")
        self.textField3.setText("")
        self.inputIndicator.clear()

        g_histories.clear()
        g_userData.clear()

    def loginWidgetNavHandler(self):
        global g_histories
        global g_date
        global g_first_counter
        global g_timestamp
        global g_userData

        now = QTime.currentTime().toString("hh:mm:ss  /  ")
        g_userData = [
            self.textField1.text(),
            self.textField2.text(),
            self.textField3.text()
            ]

        # 1. 입력 칸이 모두 채워져 있는지?
        if len(self.textField1.text()) < 1 or len(self.textField2.text()) < 1 or len(self.textField3.text()) < 1:
            self.inputIndicator.setText(now + "모든 내용을 입력해주세요.")
            return False
        
        # 2. 학번의 형태가 정상적인지?
        if (len(self.textField1.text()) == 8 and not self.textField1.text().isdigit()) or len(self.textField1.text()) != 8:
            self.inputIndicator.setText(now + "학번은 8자리 숫자입니다.")
            return False

        temp = data.isUserInDatabase(g_wb_data, self.textField2.text().strip(), self.textField1.text().strip())
        # 3. 학번과 이름 둘 중 하나만 DB에 있다면:
        if temp == 1:
            self.inputIndicator.setText(now + "이름 또는 학번을 다시 한 번 확인해주세요.")
            return False
        
        # 4. 학번과 이름이 둘 다 DB에 없다면:
        if temp == 2:
            self.inputIndicator.setText(now + "데이터베이스에서 학우님의 정보를 확인할 수 없습니다.")
            return False

        g_histories = data.getUserHistories(g_wb_data, g_userData)
        # 5. 해당 동아리 가입 이력이 없다면:
        if not len(g_histories):
            self.inputIndicator.setText(now + f"학우님의 {g_userData[2]} 동아리 활동 이력을 확인할 수 없습니다. 다른 동아리로 시도해보세요.")
            return False

        # 유저 정보 업데이트 후 다음 페이지로 진행
        g_userData = data.getUserData(g_wb_data, g_userData, g_clubs)

        # 6. 생일이 이상하다면:
        if g_userData[4] == False:
            tbirth_y = self.textField4.text()[:4]
            tbirth_m = self.textField4.text()[4:6]
            tbirth_d = self.textField4.text()[6:]

            if tbirth_m[0] == "0": tbirth_m = tbirth_m[1:]          
            if tbirth_d[0] == "0": tbirth_d = tbirth_d[1:]   
            g_userData[4] = f"{tbirth_y}년 {tbirth_m}월 {tbirth_d}일"

        mainPage.setCurrentIndex(mainPage.currentIndex() + 1)
        reviewWidget.initUI()
        print(g_userData)
        return True
    
    def initUI(self):
        global g_histories
        global g_date
        global g_first_counter
        global g_timestamp
        global g_userData

        vBox = QVBoxLayout()

        titleLabel1 = QLabel(self)
        titleLabel1.setText("활동 인증서")
        titleLabel1.setFont(font_appTitle1)
        titleLabel1.setAlignment(Qt.AlignCenter)
        titleLabel1.setStyleSheet("color: #FFB10A;")
        titleLabel2 = QLabel(self)
        titleLabel2.setText("발급 서비스")
        titleLabel2.setFont(font_appTitle2)
        titleLabel2.setAlignment(Qt.AlignCenter)
        titleLabel2.setStyleSheet("color: #332D24;")

        textFieldLabel1 = QLabel(self)
        textFieldLabel1.setText("학번")
        textFieldLabel1.setFont(font_textLabel)
        textFieldLabel1.setStyleSheet("color: #332D24;")
        textFieldLabel2 = QLabel(self)
        textFieldLabel2.setText("이름")
        textFieldLabel2.setFont(font_textLabel)
        textFieldLabel2.setStyleSheet("color: #332D24;")
        textFieldLabel3 = QLabel(self)
        textFieldLabel3.setText("동아리명")
        textFieldLabel3.setFont(font_textLabel)
        textFieldLabel3.setStyleSheet("color: #332D24;")
        textFieldLabel4 = QLabel(self)
        textFieldLabel4.setText("생일  (예시:  19990101)")
        textFieldLabel4.setFont(font_textLabel)
        textFieldLabel4.setStyleSheet("color: #332D24;")

        self.inputIndicator = QLabel()
        self.inputIndicator.setText("")
        self.inputIndicator.setFont(font_error)
        self.inputIndicator.setAlignment(Qt.AlignCenter)
        self.inputIndicator.setStyleSheet("color: #FF3C00;")
        self.inputIndicator.setWordWrap(True)

        self.textField1 = QLineEdit()
        self.textField1.setFixedSize(300, 40)
        self.textField1.setFont(font_textValue)
        self.textField1.setStyleSheet(
            "border-radius: 10px;"
            "border-style: solid;"
            "border-width: 1px;"
            "border-color: #332D24;"
        )
        self.textField2 = QLineEdit()
        self.textField2.setFixedSize(300, 40)
        self.textField2.setFont(font_textValue)
        self.textField2.setStyleSheet(
            "border-radius: 10px;"
            "border-style: solid;"
            "border-width: 1px;"
            "border-color: #332D24;"
        )
        self.textField3 = QLineEdit()
        self.textField3.setFixedSize(300, 40)
        self.textField3.setFont(font_textValue)
        self.textField3.setStyleSheet(
            "border-radius: 10px;"
            "border-style: solid;"
            "border-width: 1px;"
            "border-color: #332D24;"
        )
        self.textField4 = QLineEdit()
        self.textField4.setFixedSize(300, 40)
        self.textField4.setFont(font_textValue)
        self.textField4.setStyleSheet(
            "border-radius: 10px;"
            "border-style: solid;"
            "border-width: 1px;"
            "border-color: #332D24;"
        )

        btn = QPushButton("발급 이력 확인", self)
        btn.setStyleSheet(
            "border-radius: 15px;"
            "background-color: #E8B602;"
        )
        btn.setFont(font_btn)
        btn.setFixedHeight(50)
        btn.clicked.connect(self.loginWidgetNavHandler)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(10)
        shadow.setOffset(0, 2)
        btn.setGraphicsEffect(shadow)
        btn.setShortcut("Return")

        # 디버그 코드
        # self.textField1.setText("20185456")
        # self.textField2.setText("강시운")
        # self.textField3.setText("비꼼")

        vBox.setAlignment(Qt.AlignCenter)
        vBox.addWidget(titleLabel1)
        vBox.addWidget(titleLabel2)
        vBox.addSpacing(20)
        vBox.addWidget(textFieldLabel1)
        vBox.addWidget(self.textField1)
        vBox.addSpacing(20)
        vBox.addWidget(textFieldLabel2)
        vBox.addWidget(self.textField2)
        vBox.addSpacing(20)
        vBox.addWidget(textFieldLabel3)
        vBox.addWidget(self.textField3)
        vBox.addSpacing(20)
        vBox.addWidget(textFieldLabel4)
        vBox.addWidget(self.textField4)
        vBox.addSpacing(20)
        vBox.addWidget(self.inputIndicator)
        vBox.addSpacing(20)
        vBox.addWidget(btn)

        self.setLayout(vBox)

class ReviewWidget(QWidget):
    def __init__(self):
        super().__init__()
    
    def reviewWidgetNavHandler(self):
        mainPage.setCurrentIndex(mainPage.currentIndex() + 1)
        loadingWidget.initUI()

    def initUI(self):
        global g_histories
        global g_date
        global g_first_counter
        global g_timestamp
        global g_userData

        qApp.processEvents()
        vBox = QVBoxLayout()

        self.boxTitle1 = QLabel()
        self.boxTitle2 = QLabel()
        self.boxTitle3 = QLabel()

        self.boxTitle1.setText(f"{g_userData[1]} 회원님의")
        self.boxTitle2.setText(f"{g_userData[2]} 동아리")
        self.boxTitle3.setText("활동 내역이에요.")
        self.boxTitle1.setFont(font_boxTitle)
        self.boxTitle2.setFont(font_boxTitle)
        self.boxTitle3.setFont(font_boxTitle)
        self.boxTitle1.setAlignment(Qt.AlignLeft)
        self.boxTitle2.setAlignment(Qt.AlignLeft)
        self.boxTitle3.setAlignment(Qt.AlignLeft)
        self.boxTitle1.setFixedWidth(400)
        self.boxTitle1.setStyleSheet("color: #FFB10A;")
        self.boxTitle1.repaint()
        self.boxTitle2.repaint()
        self.boxTitle3.repaint()

        dataLabel = QLabel()
        dataString = str()
        for history in g_histories:
            tempStr1 = history[2] if history[2] else "일반 회원"
            tempStr2 = f" - {history[0]}년 {history[1]}학기  /  '{tempStr1}'으로 활동\n"
            dataString += tempStr2
        tempStr2 = tempStr2[:-1]
        dataLabel.setText(dataString)
        dataLabel.setFont(font_boxComment)

        dataArea = QScrollArea()        
        dataArea.setWidget(dataLabel)
        dataArea.setFixedWidth(400)

        btn = QPushButton("인증서 발급하기")
        btn.setStyleSheet(
            "border-radius: 15px;"
            "background-color: #E8B602;"
        )
        btn.setFont(font_btn)
        btn.setFixedHeight(50)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(10)
        shadow.setOffset(0, 2)
        btn.setGraphicsEffect(shadow)
        btn.clicked.connect(self.reviewWidgetNavHandler)

        vBox.setAlignment(Qt.AlignCenter)
        vBox.addSpacing(40)
        vBox.addWidget(self.boxTitle1)
        vBox.addWidget(self.boxTitle2)
        vBox.addWidget(self.boxTitle3)
        vBox.addSpacing(20)
        vBox.addWidget(dataArea)
        vBox.addSpacing(20)
        vBox.addWidget(btn)
        vBox.addSpacing(40)
        
        self.setLayout(vBox)

class LoadingWidget(QWidget):
    def __init__(self):
        super().__init__()

    def loadingWidgetNavHandler(self):
        global g_histories
        global g_date
        global g_first_counter
        global g_timestamp
        global g_userData
        global g_wb_data
        global g_wb_log
        global g_doc
        global g_wb_clubs
        global g_clubs
        global g_fileName

        try:
            g_fileName = data.issueDoc(g_wb_data, g_wb_log, g_doc, g_histories, g_timestamp, g_userData)
            if g_fileName:
                filePath = "./data/issued/" + g_fileName
                printer = QPrinter(QPrinter.HighResolution)
                dialog = QPrintDialog(printer, self)
                if dialog.exec_() == QDialog.Accepted:
                    with tempfile.TemporaryDirectory() as path:
                        images = convert_from_path(filePath, dpi = 300, output_folder = path)
                        painter = QPainter()
                        painter.begin(printer)
                        for i, image in enumerate(images):
                            if i > 0:
                                printer.newPage()
                            rect = painter.viewport()
                            qtImage = ImageQt(image)
                            qtImageScaled = qtImage.scaled(rect.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
                            painter.drawImage(rect, qtImageScaled)
                        painter.end()
                mainPage.setCurrentIndex(mainPage.currentIndex() + 1)
                endWidget.initUI()
            else:
                mainPage.setCurrentIndex(mainPage.currentIndex() + 2)
                errorWidget.initUI()
        except Exception as e:
            print(e)
            mainPage.setCurrentIndex(mainPage.currentIndex() + 2)
            errorWidget.initUI()

    def initUI(self):
        vBox = QVBoxLayout()
        
        boxTitle1 = QLabel(self)
        boxTitle2 = QLabel(self)
        boxTitle1.setText("인증서를\n발급하고 있습니다.")
        boxTitle2.setText("잠시만 기다려주세요.")
        boxTitle1.setFont(font_boxTitle)
        boxTitle2.setFont(font_boxSubtitle)
        boxTitle1.setAlignment(Qt.AlignCenter)
        boxTitle2.setAlignment(Qt.AlignCenter)
        boxTitle1.setFixedWidth(400)
        boxTitle1.setStyleSheet("color: #FFB10A;")

        vBox.setAlignment(Qt.AlignCenter)
        vBox.addWidget(boxTitle1)
        vBox.addSpacing(20)
        vBox.addWidget(boxTitle2)

        self.setLayout(vBox)
        qApp.processEvents()
        self.loadingWidgetNavHandler()

class EndWidget(QWidget):
    def __init__(self):
        super().__init__()

    def endWidgetNavHandler(self):
        sys.exit()

    def initUI(self):
        vBox = QVBoxLayout()

        boxTitle1 = QLabel(self)
        boxTitle2 = QLabel(self)
        boxTitle1.setText("인증서 발급이")
        boxTitle2.setText("완료되었어요!")
        boxTitle1.setFont(font_boxTitle)
        boxTitle2.setFont(font_boxTitle)
        boxTitle1.setAlignment(Qt.AlignCenter)
        boxTitle2.setAlignment(Qt.AlignCenter)
        boxTitle1.setFixedWidth(400)
        boxTitle1.setStyleSheet("color: #FFB10A;")

        boxComment = QLabel(self)
        boxComment.setText("아래 버튼을 눌러\n프로그램을 종료하세요.")
        boxComment.setWordWrap(True)
        boxComment.setAlignment(Qt.AlignCenter)
        boxComment.setFont(font_boxSubtitle)

        btn = QPushButton("프로그램 종료", self)
        btn.setStyleSheet(
            "border-radius: 15px;"
            "background-color: #E8B602;"
        )
        btn.setFont(font_btn)
        btn.setFixedHeight(50)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(10)
        shadow.setOffset(0, 2)
        btn.setGraphicsEffect(shadow)
        btn.clicked.connect(self.endWidgetNavHandler)

        vBox.setAlignment(Qt.AlignCenter)
        vBox.addSpacing(40)
        vBox.addWidget(boxTitle1)
        vBox.addWidget(boxTitle2)
        vBox.addSpacing(10)
        vBox.addWidget(boxComment)
        vBox.addSpacing(60)
        vBox.addWidget(btn)
        vBox.addSpacing(40)

        self.setLayout(vBox)

class ErrorWidget(QWidget):
    def __init__(self):
        super().__init__()
    
    def errorWidgetNavHandler(self):
        sys.exit()

    def initUI(self):
        vBox = QVBoxLayout()

        boxTitle1 = QLabel(self)
        boxTitle2 = QLabel(self)
        boxTitle1.setText("오류가")
        boxTitle2.setText("발생했어요 ㅠ_ㅠ")
        boxTitle1.setFont(font_boxTitle)
        boxTitle2.setFont(font_boxTitle)
        boxTitle1.setAlignment(Qt.AlignCenter)
        boxTitle2.setAlignment(Qt.AlignCenter)
        boxTitle1.setFixedWidth(400)
        boxTitle1.setStyleSheet("color: #FFB10A;")

        boxComment = QLabel(self)
        boxComment.setText("아래 버튼을 눌러\n프로그램을 종료하세요.")
        boxComment.setWordWrap(True)
        boxComment.setAlignment(Qt.AlignCenter)
        boxComment.setFont(font_boxSubtitle)

        btn = QPushButton("프로그램 종료", self)
        btn.setStyleSheet(
            "border-radius: 15px;"
            "background-color: #E8B602;"
        )
        btn.setFont(font_btn)
        btn.setFixedHeight(50)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(10)
        shadow.setOffset(0, 2)
        btn.setGraphicsEffect(shadow)
        btn.clicked.connect(self.errorWidgetNavHandler)

        vBox.setAlignment(Qt.AlignCenter)
        vBox.addSpacing(40)
        vBox.addWidget(boxTitle1)
        vBox.addWidget(boxTitle2)
        vBox.addSpacing(10)
        vBox.addWidget(boxComment)
        vBox.addSpacing(60)
        vBox.addWidget(btn)
        vBox.addSpacing(40)

        self.setLayout(vBox)

if __name__ == "__main__":
    app = QApplication(sys.argv)

    global g_wb_data
    global g_wb_log
    global g_doc
    global g_wb_clubs

    try:
        # 데이터 이니셜라이징
        g_wb_data = openpyxl.load_workbook(filename = "./data/data.xlsx", data_only = True)
        g_wb_log = openpyxl.load_workbook(filename = "./data/log.xlsx", data_only = True)
        g_doc = docx.Document("./data/origin_not_stamped.docx")
        g_wb_clubs = openpyxl.load_workbook(filename = "./data/clubs.xlsx", data_only = True)
    except:
        print("failed to load data")
        exit(-1)

    # 변수 선언
    global g_histories
    global g_date
    global g_first_counter
    global g_timestamp
    global g_userData
    global g_fileName
    global g_clubs
    g_histories = []
    g_date = ""
    g_first_counter = False
    g_timestamp = time.localtime()
    g_userData = []
    g_fileName = ""
    g_clubs = dict()

    for i in range(2, g_wb_clubs["List"].max_row + 1):
        g_clubs[g_wb_clubs["List"]["A" + str(i)].value] = g_wb_clubs["List"]["B" + str(i)].value


    # 폰트 초기화
    fontDb = QFontDatabase()
    fontDb.addApplicationFont("assets/NanumR.ttf")
    fontDb.addApplicationFont("assets/NanumB.ttf")
    fontDb.addApplicationFont("assets/NanumEB.ttf")
    font_appTitle1 = QFont("NanumSquare", 25, QFont.ExtraBold)
    font_appTitle2 = QFont("NanumSquare", 25, QFont.ExtraBold)
    font_textLabel = QFont("NanumSquare", 10, QFont.Bold)
    font_btn = QFont("NanumSquare", 15, QFont.Bold)
    font_boxTitle = QFont("NanumSquare", 20, QFont.Bold)
    font_boxSubtitle = QFont("NanumSquare", 15)
    font_boxComment = QFont("NanumSquare", 12)
    font_textValue = QFont("NanumSquare", 12)
    font_error = QFont("NanumSquare", 12)
    font_appTitle1.setLetterSpacing(QFont.AbsoluteSpacing, -2)
    font_appTitle2.setLetterSpacing(QFont.AbsoluteSpacing, -2)
    font_textLabel.setLetterSpacing(QFont.AbsoluteSpacing, -2)
    font_btn.setLetterSpacing(QFont.AbsoluteSpacing, -2)
    font_boxTitle.setLetterSpacing(QFont.AbsoluteSpacing, -2)
    font_boxSubtitle.setLetterSpacing(QFont.AbsoluteSpacing, -2)
    font_boxComment.setLetterSpacing(QFont.AbsoluteSpacing, -2)
    font_textValue.setLetterSpacing(QFont.AbsoluteSpacing, -2)
    font_error.setLetterSpacing(QFont.AbsoluteSpacing, -2)

    # 메인 위젯 생성
    mainPage = QStackedWidget()
    mainPage.setFixedSize(480, 640)
    mainPage.setStyleSheet(
        "background-color: #EFE0CF;"
        "border-radius: 30px;"
    )
    shadow = QGraphicsDropShadowEffect()
    shadow.setBlurRadius(50)
    shadow.setOffset(0, 2)
    mainPage.setGraphicsEffect(shadow)

    # 백그라운드 페이지 생성
    backgroundPage = QWidget()
    backgroundLayout = QVBoxLayout()
    backgroundLayout.addSpacing(3)
    backgroundLayout.addWidget(mainPage)
    backgroundLayout.addSpacing(3)
    backgroundLayout.setAlignment(Qt.AlignCenter)
    #backgroundLayout.setContentsMargins(100, 100, 100, 30)
    backgroundPage.setLayout(backgroundLayout)
    backgroundPage.setStyleSheet(
        "background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #E88802, stop:1 #E8B602)"
        )
    backgroundPage.setWindowTitle("활동 인증서 발급 서비스 - 중앙대학교 서울캠퍼스 동아리연합회")
    backgroundPage.setWindowIcon(QIcon("./assets/logo.png"))

    # 페이지 위젯을 메인 위젯에 탑재
    loginWidget = LoginWidget()
    reviewWidget = ReviewWidget()
    loadingWidget = LoadingWidget()
    endWidget = EndWidget()
    errorWidget = ErrorWidget()
    mainPage.addWidget(loginWidget)
    mainPage.addWidget(reviewWidget)
    mainPage.addWidget(loadingWidget)
    mainPage.addWidget(endWidget)
    mainPage.addWidget(errorWidget)

    # 창 위치를 디스플레이 중앙으로 이동
    currWindowSize = backgroundPage.frameGeometry()
    displayCenterPos = QDesktopWidget().availableGeometry().center()
    currWindowSize.moveCenter(displayCenterPos)
    backgroundPage.move(currWindowSize.topLeft())
    backgroundPage.show()
    sys.exit(app.exec_())